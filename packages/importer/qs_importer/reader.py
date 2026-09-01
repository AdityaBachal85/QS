"""Two-pass staging reader.

``openpyxl`` gives you either the formulas or the cached values, never both in
one pass.  Both are needed: the formula says what was *meant*, the cached value
says what Excel actually produced, and the interesting defects live in the gap
between them (``Electrical!G4 = G104*0`` has a perfectly good formula and a
cached value of zero).

Nothing is interpreted here.  This is the raw substrate every mapper reads
from, and it keeps the source cell address against every value so that any
figure in the platform can be traced back to ``Flat Sizes!D12``.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class StagingCell:
    sheet: str
    ref: str
    row: int
    col: int
    formula: str | None
    value: object

    @property
    def has_formula(self) -> bool:
        return isinstance(self.formula, str) and self.formula.startswith("=")

    @property
    def is_populated(self) -> bool:
        return self.formula is not None or self.value is not None

    @property
    def full_ref(self) -> str:
        return f"{self.sheet}!{self.ref}"

    def as_float(self, default: float | None = None) -> float | None:
        try:
            return float(self.value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return default

    def as_text(self, default: str = "") -> str:
        if self.value is None:
            return default
        return str(self.value).strip()


class Workbook:
    """A staged workbook: every populated cell, formula and value together."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self.sheets: list[str] = []
        self._cells: dict[str, dict[str, StagingCell]] = {}
        self._load()

    def _load(self) -> None:
        # Not read_only: that mode yields EmptyCell placeholders with no column
        # index, which makes aligning the two passes fragile. These workbooks are
        # ~1 MB, so holding both in memory is cheaper than the alignment bugs.
        formulas = load_workbook(self.path, data_only=False)
        values = load_workbook(self.path, data_only=True)
        try:
            self.sheets = list(formulas.sheetnames)
            for name in self.sheets:
                self._cells[name] = self._stage_sheet(formulas[name], values[name], name)
        finally:
            formulas.close()
            values.close()

    @staticmethod
    def _stage_sheet(f_sheet, v_sheet, name: str) -> dict[str, StagingCell]:
        staged: dict[str, StagingCell] = {}
        for f_row in f_sheet.iter_rows():
            for cell in f_row:
                formula = cell.value
                value = v_sheet.cell(row=cell.row, column=cell.column).value
                if formula is None and value is None:
                    continue
                is_formula = isinstance(formula, str) and formula.startswith("=")
                ref = f"{get_column_letter(cell.column)}{cell.row}"
                staged[ref] = StagingCell(
                    sheet=name, ref=ref, row=cell.row, col=cell.column,
                    formula=formula if is_formula else None,
                    value=value if is_formula else formula,
                )
        return staged

    # -- access ------------------------------------------------------------

    def sheet(self, name: str) -> dict[str, StagingCell]:
        if name not in self._cells:
            raise KeyError(f"no sheet named {name!r} in {self.path.name}")
        return self._cells[name]

    def cell(self, sheet: str, ref: str) -> StagingCell:
        return self.sheet(sheet).get(
            ref, StagingCell(sheet, ref, 0, 0, None, None)
        )

    def value(self, sheet: str, ref: str, default: object = None) -> object:
        cell = self.cell(sheet, ref)
        return cell.value if cell.value is not None else default

    def number(self, sheet: str, ref: str, default: float | None = None) -> float | None:
        return self.cell(sheet, ref).as_float(default)

    def text(self, sheet: str, ref: str, default: str = "") -> str:
        return self.cell(sheet, ref).as_text(default)

    def formula(self, sheet: str, ref: str) -> str | None:
        return self.cell(sheet, ref).formula

    def rows(self, sheet: str) -> Iterator[int]:
        seen = {c.row for c in self.sheet(sheet).values()}
        return iter(sorted(seen))

    def census(self) -> dict[str, tuple[int, int]]:
        """Per sheet: (populated cells, formula cells).  The Stage 0 gate."""
        return {
            name: (len(cells), sum(1 for c in cells.values() if c.has_formula))
            for name, cells in self._cells.items()
        }

    def totals(self) -> tuple[int, int, int]:
        census = self.census()
        return (len(census),
                sum(p for p, _ in census.values()),
                sum(f for _, f in census.values()))
