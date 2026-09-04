"""The read model: everything the UI displays, computed from the engine.

Every figure here is produced by calling into ``qs_engine``. Nothing is cached in
the database, nothing is recomputed by the UI, and there is no second
implementation of any rule. That is what makes the automation real: change one
input and every view that depends on it moves, because they all ask the engine
the same question again.
"""

from __future__ import annotations

import re
from typing import Any, Iterable

from qs_engine.model import (OpeningKind, Project, ProjectModel,
                             RoomCategory)
from qs_engine.params import ParameterSet
from qs_engine.rules.rate_buildup import RateBuildupError, effective_rate
from qs_engine.rules.room_qty import (RULE_DEDUCTIONS, NegativeNetQuantityError,
                                      compute_room_quantity)
from qs_engine.rules.schedule import (opening_schedule, opening_totals,
                                      priced_opening_schedule,
                                      total_opening_amount, total_openings)
from qs_engine.rules.unit_area import (room_area_sqft, room_total_sqft,
                                       unit_type_area_sqft,
                                       unit_type_area_sqm, unit_type_total_sqft)
from qs_engine.units import UnitMismatchError
from qs_engine.validation import Severity, validate


def _derivation(derived) -> dict[str, Any]:
    d = derived.derivation
    return {
        "rule": d.rule,
        "expression": d.expression,
        "inputs": [{"name": i.name, "value": i.value, "source": i.source}
                   for i in d.inputs],
        "excel_ref": d.excel_ref,
        "note": d.note,
    }


def headline(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """The numbers in the top bar. Recomputed on every request, deliberately."""
    flats = sum(model.unit_count(u.id) for u in model.unit_types
                if u.is_residential and not u.is_common_area)
    offices = sum(model.unit_count(u.id) for u in model.unit_types
                  if u.classification == "Office")
    carpet = sum(unit_type_area_sqft(u.id, model, params).value * model.unit_count(u.id)
                 for u in model.unit_types if not u.is_common_area)
    report = validate(model, params)
    return {
        "project": {"id": model.project.id, "code": model.project.code,
                    "name": model.project.name, "city": model.project.city},
        "floors": len(model.floors),
        "flats": flats,
        "offices": offices,
        "classification": model.counts_by_classification(),
        "building_height_m": round(sum(f.floor_to_floor_ht for f in model.floors), 2),
        "carpet_area_sqft": carpet,
        "doors": total_openings(model, (OpeningKind.DOOR,)).value,
        "rooms": len(model.unit_type_rooms),
        "rate_items": len(model.rate_items),
        "health": {
            "score": round(report.health_score()),
            "blocking": len(report.blocking),
            "warnings": len(report.of(Severity.WARNING)),
            "info": len(report.of(Severity.INFO)),
            "can_issue": report.can_issue,
        },
    }


def room_config(model: ProjectModel) -> dict[str, Any]:
    """The floor x unit-type matrix, plus its live totals."""
    types = sorted(model.unit_types, key=lambda u: u.seq)
    mix = {(m.floor_id, m.unit_type_id): m.count for m in model.floor_unit_mix}
    return {
        "unit_types": [
            {"id": u.id, "code": u.code, "classification": u.classification,
             "is_common_area": u.is_common_area,
             "total": model.unit_count(u.id),
             "count_override": u.count_override}
            for u in types
        ],
        "floors": [
            {"id": f.id, "seq": f.seq, "name": f.name,
             "floor_to_floor_ht": f.floor_to_floor_ht,
             "floor_type": f.floor_type.value,
             "counts": {u.id: mix.get((f.id, u.id), 0) for u in types},
             "row_total": sum(mix.get((f.id, u.id), 0) for u in types
                              if not u.is_common_area),
             "row_total_derivation": _floor_total(f, types, mix)}
            for f in sorted(model.floors, key=lambda f: f.seq)
        ],
        "classification": model.counts_by_classification(),
    }


def _floor_total(floor, types, mix) -> dict[str, Any]:
    """The units on one floor: a fold across the row, common areas excluded.

    Common areas are excluded and *named* as excluded rather than silently
    dropped -- a lobby is on the floor, it is just not a unit anybody buys.
    """
    from qs_engine.provenance import Input, derive

    inputs = [Input(u.code, mix.get((floor.id, u.id), 0),
                    "counted from the floor matrix")
              for u in types
              if not u.is_common_area and mix.get((floor.id, u.id), 0)]
    excluded = [u.code for u in types
                if u.is_common_area and mix.get((floor.id, u.id), 0)]
    total = sum(i.value for i in inputs)
    return _derivation(derive(
        total, "units_on_floor",
        " + ".join(f"{i.value:g}" for i in inputs) or "0", inputs,
        note=("Common areas on this floor are not counted as units: "
              + ", ".join(excluded) + ". They keep their count and are "
              "excluded by name, never by multiplying by zero."
              if excluded else
              "Every unit type on this floor, added. The BHK split elsewhere "
              "is a group-by over this same matrix, so it cannot go stale the "
              "way Room Conf!L44 did (C-21).")))


def unit_types(model: ProjectModel, params: ParameterSet) -> list[dict[str, Any]]:
    out = []
    for u in sorted(model.unit_types, key=lambda x: x.seq):
        rooms = model.rooms_of(u.id)
        per_unit = unit_type_area_sqft(u.id, model, params)
        total = unit_type_total_sqft(u.id, model, params)
        out.append({
            "id": u.id, "code": u.code, "classification": u.classification,
            "is_common_area": u.is_common_area,
            "rooms": len(rooms),
            # How many rooms in this type are measured off a counter, so the
            # list can say where the Kitchen platforms tab will be. Not every
            # flat has a kitchen -- Flat 1A does not -- and hunting for a tab
            # on a unit type that cannot have one is a bad way to find that out.
            "counter_rooms": len(counter_rooms(model, u.id)),
            "count": model.unit_count(u.id),
            "area_sqm": unit_type_area_sqm(u.id, model).value,
            "area_sqft": per_unit.value,
            "total_sqft": total.value,
            "derivation": _derivation(total),
        })
    return out


def unit_rooms(model: ProjectModel, params: ParameterSet,
               unit_type_id: str) -> dict[str, Any]:
    """One unit type's rooms, with every derived figure and its openings."""
    unit = model.unit_type(unit_type_id)

    # Wall quantities depend on the floor the unit sits on, so this view has to
    # resolve the same height the take-off does -- otherwise the two screens
    # would report different walls for one room, which is the class of defect
    # this platform exists to remove. Where a type spans floors of different
    # heights the rooms are shown at the predominant one, and every height is
    # reported alongside so the split is visible rather than averaged away.
    placements = model.height_placements(unit_type_id)
    dominant = max(placements, key=lambda p: p.count) if placements else None
    floor_height_m = dominant.height_m if dominant else None

    openings_by_room: dict[str, list[dict[str, Any]]] = {}
    for o in model.room_openings:
        ot = model.opening_type(o.opening_type_id)
        openings_by_room.setdefault(o.unit_type_room_id, []).append({
            "id": o.id, "code": ot.code, "kind": ot.kind.value,
            "count": o.count, "width_m": ot.width_m, "height_m": ot.height_m,
            "opening_type_id": ot.id,
        })

    rows = []
    for r in model.rooms_of(unit_type_id):
        area = room_area_sqft(r, params)
        rows.append({
            "id": r.id, "seq": r.seq, "label": r.label,
            "room_type_id": r.room_type_id,
            "room_type": model.room_type(r.room_type_id).name,
            "category": model.room_type(r.room_type_id).category.value,
            "count_per_unit": r.count_per_unit,
            "carpet_area_sqm": r.carpet_area_sqm,
            "perimeter_m": r.perimeter_m,
            "clear_height_m": r.clear_height_m,
            "area_sqft": area.value,
            "total_sqft": room_total_sqft(r, params).value,
            "derivation": _derivation(area),
            "openings": openings_by_room.get(r.id, []),
            "quantities": room_quantities(model, params, r.id, floor_height_m),
        })
    return {
        "unit_type": {"id": unit.id, "code": unit.code,
                      "classification": unit.classification,
                      "count": model.unit_count(unit.id)},
        "floor_height_m": floor_height_m,
        "heights": [{"height_m": p.height_m, "count": p.count,
                     "floors": list(p.floors)} for p in placements],
        "rooms": rows,
        "area_sqft": unit_type_area_sqft(unit_type_id, model, params).value,
        "total_sqft": unit_type_total_sqft(unit_type_id, model, params).value,
    }


#: The finishes shown per room, in the order a QS reads them.
QUANTITY_RULES = ("floor_area", "skirting", "wall_finish", "dado",
                  "ceiling_area", "door_frame", "window_frame")


def room_quantities(model: ProjectModel, params: ParameterSet,
                    room_id: str,
                    floor_height_m: float | None = None) -> list[dict[str, Any]]:
    """Gross, deduction and net for every finish in one room.

    This is the answer to "when we give costing we minus the doors and windows,
    then skirting". The deduction is a fold over the room's own openings, so
    adding a door here changes it with no linking step -- and a deduction that
    does not match the quantity's unit raises rather than computing (C-35).
    """
    room = model.room(room_id)
    out = []
    for rule in QUANTITY_RULES:
        entry: dict[str, Any] = {"rule": rule,
                                 "deduction_rule": RULE_DEDUCTIONS.get(rule, "none")}
        try:
            q = compute_room_quantity(room, rule, model, params,
                                      floor_height_m=floor_height_m)
            entry.update({
                "unit": q.gross.unit.code,
                "gross": q.gross.value,
                "deduction": q.deduction.value,
                "net": q.net.value,
                "gross_derivation": _derivation(q.gross_derivation),
                "deduction_derivation": _derivation(q.deduction_derivation),
                "error": None,
            })
        except (UnitMismatchError, NegativeNetQuantityError) as exc:
            entry.update({"unit": "", "gross": None, "deduction": None,
                          "net": None, "error": str(exc)})
        out.append(entry)
    return out


#: Rules that measure off a counter rather than round the perimeter.
COUNTER_RULES = frozenset({"kitchen_platform", "service_platform"})


def counter_rooms(model: ProjectModel, unit_type_id: str) -> list:
    """The rooms in one unit type that the Kitchen platforms tab covers.

    One rule, so the list and the tab cannot disagree: a room qualifies if its
    schedule prices a counter, if it already has counters entered, or if it is
    a kitchen at all. The first is the real test -- the office Pantry is priced
    for a service counter and is not categorised as a kitchen anywhere -- and
    the last is there because you asked for the tab whenever there is a
    kitchen, priced or not.
    """
    slots = {s.id: s for s in model.finish_slots}

    def priced_for_a_counter(room_type_id: str) -> bool:
        for candidate in {room_type_id, model.pricing_room_type(room_type_id)}:
            for spec in model.room_finish_specs:
                if spec.room_type_id != candidate or not spec.is_applicable:
                    continue
                slot = slots.get(spec.finish_slot_id)
                rule = spec.qty_rule or (slot.qty_rule if slot else "")
                if rule in COUNTER_RULES:
                    return True
        return False

    out = []
    for room in model.rooms_of(unit_type_id):
        if (model.room_type(room.room_type_id).category is RoomCategory.KITCHEN
                or model.kitchen_platform(room.id) is not None
                or priced_for_a_counter(room.room_type_id)):
            out.append(room)
    return out


def kitchen_platforms(model: ProjectModel, params: ParameterSet,
                      unit_type_id: str) -> dict[str, Any]:
    """The counters in one unit type's rooms, and what they measure.

    One row per room that is priced for a counter, whether or not its runs have
    been entered yet.  A room with no counters shows blank rather than zero and
    says so, because a kitchen with no counters is unmeasured, not free.

    The two dado areas sit beside the four entries so the effect of typing is
    visible in place.  They are computed here by the engine, not in the
    browser: the same call the take-off makes, so the tab and the cost cannot
    disagree.
    """
    from qs_engine.rules.room_qty import (MissingKitchenPlatformError,
                                          QTY_RULES)

    slots = {s.id: s for s in model.finish_slots}
    counter_rules = {"kitchen_platform", "service_platform"}

    def measures_off_a_counter(room_type_id: str) -> set[str]:
        """The counter-driven rules this room type is priced for."""
        found = set()
        for spec in model.room_finish_specs:
            if spec.room_type_id != room_type_id or not spec.is_applicable:
                continue
            slot = slots.get(spec.finish_slot_id)
            rule = spec.qty_rule or (slot.qty_rule if slot else "")
            if rule in counter_rules or rule.startswith("dado_"):
                found.add(rule)
        return found

    rows = []
    for room in counter_rooms(model, unit_type_id):
        room_type = model.room_type(room.room_type_id)
        # Resolved through the pricing mapping too: a room type priced entirely
        # through another one carries no finish specs of its own.
        priced_for = (measures_off_a_counter(room.room_type_id)
                      | measures_off_a_counter(
                          model.pricing_room_type(room.room_type_id)))
        platform = model.kitchen_platform(room.id)
        entry: dict[str, Any] = {
            "id": platform.id if platform else None,
            "unit_type_room_id": room.id,
            "room_label": room.label or room_type.name,
            "room_type": room_type.name,
            "priced": bool(priced_for & counter_rules),
            "count_per_unit": room.count_per_unit,
            "main_platform_m": platform.main_platform_m if platform else None,
            "service_platform_m": platform.service_platform_m if platform else None,
            "dado_above_m": platform.dado_above_m if platform else None,
            "dado_below_m": platform.dado_below_m if platform else None,
            "priced_for": sorted(priced_for),
        }
        for rule, key in (("dado_above_platform", "dado_above"),
                          ("dado_below_platform", "dado_below")):
            try:
                derived = QTY_RULES[rule](room, model, params)
                entry[key] = derived.value.value
                entry[f"{key}_derivation"] = _derivation(derived)
            except MissingKitchenPlatformError as exc:
                entry[key] = None
                entry[f"{key}_derivation"] = None
                entry.setdefault("message", str(exc))
        rows.append(entry)
    return {"rooms": rows}


def _opening_area(line) -> dict[str, Any]:
    """Width x height, said out loud.

    A one-step calculation still deserves its working: it is the difference
    between "1.89" and "1.20 x 1.575, the leaf as scheduled".
    """
    from qs_engine.provenance import Input, derive
    return _derivation(derive(
        line.width_m * line.height_m, "opening_area",
        f"{line.width_m:g} x {line.height_m:g}",
        [Input("width", line.width_m, "on the opening type"),
         Input("height", line.height_m, "on the opening type")],
        note="The leaf as scheduled. What is deducted from a wall is this area "
             "times the number in the room; what is deducted from skirting is "
             "the width alone, because a running metre takes a running-metre "
             "deduction (C-35)."))


def openings(model: ProjectModel) -> dict[str, Any]:
    """The door and window schedule -- a query, not a bounded range (C-18)."""
    def lines(kinds):
        return [{"code": l.code, "kind": l.kind.value, "width_m": l.width_m,
                 "height_m": l.height_m, "count": l.count,
                 "quantity": l.quantity, "unit": l.unit,
                 "count_derivation": _derivation(l.count_derivation)
                 if l.count_derivation else None,
                 "quantity_derivation": _derivation(l.quantity_derivation)
                 if l.quantity_derivation else None,
                 "area_derivation": _opening_area(l)}
                for l in opening_schedule(model, kinds)]

    return {
        "types": [{"id": o.id, "code": o.code, "kind": o.kind.value,
                   "width_m": o.width_m, "height_m": o.height_m,
                   "area_sqm": o.area_sqm, "specification": o.specification}
                  for o in sorted(model.opening_types, key=lambda o: (o.kind.value, o.code))],
        "doors": lines((OpeningKind.DOOR,)),
        "windows": lines((OpeningKind.WINDOW, OpeningKind.VENTILATOR)),
        "railings": lines((OpeningKind.RAILING,)),
        "curtain_wall": lines((OpeningKind.CURTAIN_WALL,)),
        "total_doors": total_openings(model, (OpeningKind.DOOR,)).value,
    }


def rates(model: ProjectModel, params: ParameterSet) -> list[dict[str, Any]]:
    out = []
    for item in model.rate_items:
        revision = model.current_revision(item.id)
        row: dict[str, Any] = {
            "id": item.id, "code": item.code, "description": item.description,
            "specification": item.specification, "unit": item.unit,
            "category": item.category,
            "method": revision.method.value if revision else None,
            "basic_rate": revision.basic_rate if revision else None,
            "laying_rate": revision.laying_rate if revision else None,
            "wastage_pct": revision.wastage_pct if revision else None,
            "frame_width_m": revision.frame_width_m if revision else None,
            "is_priced": bool(revision and revision.is_priced),
            "revision_id": revision.id if revision else None,
        }
        try:
            rate = effective_rate(item, model, params)
            row["overall_rate"] = rate.value
            row["derivation"] = _derivation(rate)
        except RateBuildupError as exc:
            row["overall_rate"] = None
            row["derivation"] = None
            row["error"] = str(exc)
        out.append(row)
    return out


# --------------------------------------------------------------------------
# The finishing take-off -- where quantities meet rates
# --------------------------------------------------------------------------

def _contributors(lines) -> dict[str, list[dict[str, Any]]]:
    """What each group on a totals screen is made of.

    Every fold is a filter over the same take-off lines, so a group's working
    is simply the lines that matched it, gathered by whichever dimension the
    group does *not* already fix: a finish breaks down by unit type, a unit
    type by finish. Keyed the way each grid keys its rows, so a screen can look
    a group up by the row it drew.
    """
    out: dict[str, list[dict[str, Any]]] = {}

    def gather(key_of, breakdown_by):
        for line in lines:
            per = out.setdefault(key_of(line), [])
            label = breakdown_by(line)
            entry = next((e for e in per if e["label"] == label), None)
            if entry is None:
                entry = {"label": label, "quantity": 0.0, "amount": 0.0,
                         "unit": line.unit, "lines": 0}
                per.append(entry)
            entry["quantity"] += line.total_qty
            entry["amount"] += line.total_amount
            entry["lines"] += 1
            if entry["unit"] != line.unit:
                entry["unit"] = ""      # mixed units do not add

    gather(lambda l: l.finish_slot_id, lambda l: l.unit_type_code)
    gather(lambda l: l.room_type_id, lambda l: l.unit_type_code)
    gather(lambda l: l.unit_type_id, lambda l: l.finish_name)
    gather(lambda l: f"{l.finish_slot_id}|{l.room_type_id}",
           lambda l: l.unit_type_code)
    for per in out.values():
        per.sort(key=lambda e: -e["amount"])
    return out


def _line_json(line, *, working: bool = False) -> dict[str, Any]:
    """One take-off line.

    ``working`` carries the three derivation blocks. They are 54% of the
    payload and a QS opens one at a time, so the list omits them and the
    derivation route serves whichever is clicked.
    """
    out = {
        "unit_type_id": line.unit_type_id, "unit_type": line.unit_type_code,
        "room_id": line.room_id, "room": line.room_label,
        "finish": line.finish_name, "finish_slot_id": line.finish_slot_id,
        "rule": line.qty_rule, "unit": line.unit,
        "gross": line.gross, "deduction": line.deduction, "net": line.net,
        "unit_count": line.unit_count, "total_qty": line.total_qty,
        "rate_item_id": line.rate_item_id, "rate_description": line.rate_description,
        "rate": line.rate, "amount_per_unit": line.amount_per_unit,
        "total_amount": line.total_amount,
        "status": line.status, "message": line.message,
        "floor_height_m": line.floor_height_m, "floor_scope": line.floor_scope,
    }
    if working:
        out.update({
            "gross_derivation": _derivation(line.gross_derivation)
            if line.gross_derivation else None,
            "deduction_derivation": _derivation(line.deduction_derivation)
            if line.deduction_derivation else None,
            "rate_derivation": _derivation(line.rate_derivation)
            if line.rate_derivation else None,
        })
    return out


def takeoff_derivation(model: ProjectModel, params: ParameterSet, room_id: str,
                       finish_slot_id: str, unit_type_id: str | None = None,
                       floor_height_m: float | None = None) -> dict[str, Any] | None:
    """The working behind one figure, fetched when somebody clicks it."""
    from qs_engine.rules.takeoff import compute_takeoff

    for line in compute_takeoff(model, params, unit_type_id):
        if line.room_id != room_id or line.finish_slot_id != finish_slot_id:
            continue
        if floor_height_m is not None and line.floor_height_m != floor_height_m:
            continue
        return _line_json(line, working=True)
    return None


def _group_json(groups) -> list[dict[str, Any]]:
    return [{"key": g.key, "label": g.label, "unit": g.unit,
             "quantity": g.quantity, "quantity_sqft": g.quantity_sqft,
             "amount": g.amount, "lines": g.lines, "unpriced": g.unpriced,
             "blended_rate": g.blended_rate,
             "rate_per_sqft": g.rate_per_sqft} for g in groups]


def takeoff(model: ProjectModel, params: ParameterSet,
            unit_type_id: str | None = None) -> dict[str, Any]:
    """Every finish in every room, priced.

    Replaces 1,451 hand-written rows in ``Internal Finishes Flats``. The totals
    are folds over a filter, so a finish added tomorrow is included because it
    matches, not because a range was widened.
    """
    from qs_engine.rules.takeoff import (by_finish, by_unit_type, compute_takeoff,
                                         total_amount, unpriced)

    lines = compute_takeoff(model, params, unit_type_id)
    missing = unpriced(lines)
    return {
        "lines": [_line_json(l) for l in lines],
        "by_finish": _group_json(by_finish(lines, params)),
        "by_unit_type": _group_json(by_unit_type(lines, params)),
        "contributors": _contributors(lines),
        "total": total_amount(lines),
        "line_count": len(lines),
        "priced_count": sum(1 for l in lines if l.is_priced),
        "unpriced": [_line_json(l) for l in missing],
        "unpriced_qty": sum(l.total_qty for l in missing),
    }


def room_costs(model: ProjectModel, params: ParameterSet,
               unit_type_id: str) -> dict[str, list[dict[str, Any]]]:
    """Take-off lines for one unit type, keyed by room.

    This is what puts a rate beside every quantity on the Unit Types screen.
    """
    from qs_engine.rules.takeoff import compute_takeoff

    out: dict[str, list[dict[str, Any]]] = {}
    for line in compute_takeoff(model, params, unit_type_id):
        out.setdefault(line.room_id, []).append(_line_json(line, working=True))
    return out


def room_type_mapping(model: ProjectModel,
                      params: ParameterSet) -> list[dict[str, Any]]:
    """Which rate block prices each room type, and whether anyone has agreed.

    The sizes sheets and the rate list name rooms differently -- ``M. Bedroom``
    against ``M. Bed``, ``M. Toilet`` against ``Toilet With M. Bed``. Only six of
    twenty-five match by name, so without this every other room is measured and
    unpriced.
    """
    from qs_engine.rules.takeoff import by_room_type, compute_takeoff

    used = {r.room_type_id for r in model.unit_type_rooms}
    priced = {s.room_type_id for s in model.room_finish_specs}
    rooms_per_type: dict[str, int] = {}
    for room in model.unit_type_rooms:
        rooms_per_type[room.room_type_id] = rooms_per_type.get(room.room_type_id, 0) + 1

    # What each link is currently worth, so agreeing to one is an informed
    # decision rather than a shrug.
    takeoff_lines = compute_takeoff(model, params)
    worth = {g.key: g.amount for g in by_room_type(takeoff_lines)}
    #: Which unit types make up that figure, so "currently worth Rs 84 lakh"
    #: can be read back as the flats it is spread across.
    worth_from: dict[str, list[dict[str, Any]]] = {}
    for line in takeoff_lines:
        if not line.is_priced:
            continue
        entry = worth_from.setdefault(line.room_type_id, {})
        row = entry.setdefault(line.unit_type_id, {
            "label": line.unit_type_code, "amount": 0.0, "lines": 0})
        row["amount"] += line.total_amount
        row["lines"] += 1
    worth_by_unit = {k: sorted(v.values(), key=lambda r: -r["amount"])
                     for k, v in worth_from.items()}

    out = []
    for room_type in sorted(model.room_types, key=lambda t: t.name):
        if room_type.id not in used:
            continue
        target_id = model.pricing_room_type(room_type.id)
        target = model.room_type(target_id) if target_id else None
        out.append({
            "id": room_type.id, "name": room_type.name,
            "category": room_type.category.value,
            "rooms": rooms_per_type.get(room_type.id, 0),
            "prices_as_id": room_type.prices_as_id,
            "prices_as": target.name if target else "",
            "confirmed": room_type.mapping_confirmed,
            "own_schedule": room_type.id in priced,
            "finishes": len(model.finish_spec_for(room_type.id)),
            "amount": worth.get(room_type.id, 0.0),
            "worth_from": worth_by_unit.get(room_type.id, []),
        })
    return out


def priceable_room_types(model: ProjectModel) -> list[dict[str, str]]:
    """The room types that carry a finish schedule -- the dropdown for mapping."""
    priced = {s.room_type_id for s in model.room_finish_specs}
    return [{"value": t.id, "label": t.name}
            for t in sorted(model.room_types, key=lambda t: t.name)
            if t.id in priced]


def validation(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    report = validate(model, params)
    return {
        "score": round(report.health_score()),
        "summary": report.summary(),
        "can_issue": report.can_issue,
        "findings": [
            {"rule": f.rule, "severity": f.severity.value, "message": f.message,
             "entity": f.entity, "entity_id": f.entity_id, "value": f.value}
            for f in sorted(report.findings,
                            key=lambda f: {"blocking": 0, "warning": 1,
                                           "info": 2}[f.severity.value])
        ],
    }


# --------------------------------------------------------------------------
# Totals -- the whole building in one view
# --------------------------------------------------------------------------

def opening_costs(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """Every door, window, railing and curtain-wall bay, counted and priced.

    Counts were always folded correctly here; the money is new. A door is
    priced by the leaf and glazing by the square metre, and the multiplication
    goes through the unit-safe path, so a per-Nos. rate cannot be applied to an
    area.
    """
    lines = priced_opening_schedule(model, params)
    return {
        "lines": [{
            "code": l.code, "kind": l.kind.value,
            "width_m": l.width_m, "height_m": l.height_m,
            "count": l.count, "quantity": l.quantity, "unit": l.unit,
            "rate": l.rate, "rate_unit": l.rate_unit,
            "rate_item_id": l.rate_item_id,
            "rate_description": l.rate_description,
            "amount": l.amount, "status": l.status, "message": l.message,
            "count_derivation": _derivation(l.count_derivation)
            if l.count_derivation else None,
            "quantity_derivation": _derivation(l.quantity_derivation)
            if l.quantity_derivation else None,
            "rate_derivation": _derivation(l.rate_derivation)
            if l.rate_derivation else None,
            "amount_derivation": _derivation(l.amount_derivation)
            if l.amount_derivation else None,
            "area_derivation": _opening_area(l),
        } for l in lines],
        "bands": _opening_bands(model, params, lines),
        "total": total_opening_amount(model, params),
        "total_count": sum(l.count for l in lines),
        "unpriced": [{"code": l.code, "message": l.message}
                     for l in lines if not l.is_priced],
    }


def _opening_bands(model: ProjectModel, params: ParameterSet,
                   lines) -> list[dict[str, Any]]:
    """Doors, windows, railings -- each band with the types that make it up."""
    from qs_engine.provenance import Input, derive
    from qs_engine.rules.schedule import BANDS

    by_kind: dict[str, list] = {}
    for line in lines:
        by_kind.setdefault(line.kind.value, []).append(line)

    out = []
    for band in opening_totals(model, params):
        kinds = next((k for key, _, k in BANDS if key == band.key), ())
        members = [l for k in kinds for l in by_kind.get(k.value, [])]
        members.sort(key=lambda l: -l.amount)
        out.append({
            "key": band.key, "label": band.label, "unit": band.unit,
            "count": band.count, "quantity": band.quantity,
            "amount": band.amount, "lines": band.lines,
            "unpriced": band.unpriced,
            "count_derivation": _derivation(derive(
                band.count, "band_count",
                " + ".join(f"{l.count:g}" for l in members) or "0",
                [Input(l.code, l.count,
                       f"{l.width_m:g} x {l.height_m:g} m") for l in members],
                excel_ref="Doors!L141 = SUBTOTAL(9,L5:L140)",
                note="A filter over every opening type of this kind, folded up "
                     "through the rooms that carry them. The workbook's own "
                     "count and its schedule disagree -- Doors!E141 says 58 "
                     "where L141 says 2,180 -- because they were two sums over "
                     "two ranges. Here the count and the money come from one "
                     "fold, so they cannot part company (C-12).")),
            "amount_derivation": _derivation(derive(
                band.amount, "band_amount",
                " + ".join(f"{l.amount:,.0f}" for l in members if l.amount) or "0",
                [Input(l.code, l.amount,
                       (f"{l.count:g} @ {l.rate:g} per {l.rate_unit}"
                        if l.rate else l.message or "no rate"))
                 for l in members],
                note=("Every type of this kind priced on what it is bought by: "
                      "a door by the leaf, glazing by the square metre, railing "
                      "by the running metre."
                      + (f" {band.unpriced} type(s) here are measured and carry "
                         f"no price, so their quantity is real and their amount "
                         f"is missing rather than zero (C-11)."
                         if band.unpriced else "")))),
        })
    return out


def internal_finishes(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """The take-off in the shape of ``Internal Finishes``.

    The workbook lays this out one block per room, under a heading per unit
    type carrying its count, and a QS reads down it. Every other view here
    folds those lines up -- by finish, by room type, by unit type -- which
    answers different questions and none of them is "show me the take-off the
    way I read it".

    So this is the same lines in the sheet's own order:

        Flat 1A                                              10 units
          Multi Purpose Room    22.05 sq m   19.39 m      Rs 351.20 / sq ft
            Flooring        gross  deduct  net  unit  rate  per unit  total
            ...

    Computed by the engine, not read out of the workbook: the point is that it
    can be put beside the sheet and checked line for line. The Rs/sq ft against
    each room is the sheet's own column M --
    ``=SUM(L5:L16)/C4/10.764`` -- the room's whole finishing cost over its
    carpet area, in square feet.
    """
    from qs_engine.rules.takeoff import compute_takeoff, total_amount

    lines = compute_takeoff(model, params)
    per_room: dict[str, list] = {}
    for line in lines:
        per_room.setdefault(line.room_id, []).append(line)

    sqm_to_sqft = params["factor_sqm_to_sqft"]
    unit_types = []
    for unit in sorted(model.unit_types, key=lambda u: u.seq):
        rooms = []
        for room in model.rooms_of(unit.id):
            got = per_room.get(room.id, [])
            if not got:
                continue
            per_unit = sum(l.amount_per_unit for l in got if l.is_priced)
            area_sqft = room.carpet_area_sqm * sqm_to_sqft
            rooms.append({
                "id": room.id, "label": room.label,
                "room_type": model.room_type(room.room_type_id).name,
                "carpet_area_sqm": room.carpet_area_sqm,
                "perimeter_m": room.perimeter_m,
                "count_per_unit": room.count_per_unit,
                "amount_per_unit": per_unit,
                "total_amount": sum(l.total_amount for l in got if l.is_priced),
                # The sheet's column M: what this room costs to finish, per
                # square foot of its own carpet.
                "rate_per_sqft": (per_unit / area_sqft) if area_sqft else None,
                "unpriced": sum(1 for l in got if not l.is_priced),
                "lines": [_line_json(l) for l in got],
            })
        if not rooms:
            continue
        unit_types.append({
            "id": unit.id, "code": unit.code,
            "classification": unit.classification,
            "is_common_area": unit.is_common_area,
            "units": model.unit_count(unit.id),
            "rooms": rooms,
            "amount_per_unit": sum(r["amount_per_unit"] for r in rooms),
            "total_amount": sum(r["total_amount"] for r in rooms),
        })

    # The foot of the sheet: rows 1998-2038, a SUMIF per rate description.
    # A fold over the same lines rather than a second reading of them, so the
    # summary and the blocks above it cannot disagree the way the workbook's
    # two door counts do.
    summary: dict[str, dict[str, Any]] = {}
    for line in lines:
        if not line.is_priced:
            continue
        key = line.rate_description or line.finish_name
        entry = summary.setdefault(key, {
            "description": key, "unit": line.unit, "quantity": 0.0,
            "amount": 0.0, "lines": 0})
        entry["quantity"] += line.total_qty
        entry["amount"] += line.total_amount
        entry["lines"] += 1
        if entry["unit"] != line.unit:
            entry["unit"] = ""

    for entry in summary.values():
        entry["rate"] = (entry["amount"] / entry["quantity"]
                         if entry["quantity"] else None)

    return {
        "unit_types": unit_types,
        "summary": sorted(summary.values(), key=lambda e: -e["amount"]),
        "total": total_amount(lines),
        "line_count": len(lines),
        "unpriced": sum(1 for l in lines if not l.is_priced),
    }


def finish_totals(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """The building's finishing, folded three ways.

    By finish, by room type, and by the two together -- "total flooring area in
    toilets" is one cell of the last. All three are filters over the same
    take-off lines, so they cannot disagree with each other or with the
    per-room views.
    """
    from qs_engine.rules.takeoff import (by_finish, by_finish_and_room_type,
                                         by_room_type, by_unit_type,
                                         compute_takeoff, total_amount)

    lines = compute_takeoff(model, params)
    total = total_amount(lines)
    carpet = sum(unit_type_area_sqft(u.id, model, params).value * model.unit_count(u.id)
                 for u in model.unit_types if not u.is_common_area)

    contributors = _contributors(lines)

    return {
        "by_finish": _group_json(by_finish(lines, params)),
        "by_room_type": _group_json(by_room_type(lines, params)),
        "matrix": _group_json(by_finish_and_room_type(lines, params)),
        "by_unit_type": _group_json(by_unit_type(lines, params)),
        "contributors": contributors,
        "total": total,
        "openings_total": total_opening_amount(model, params),
        "carpet_area_sqft": carpet,
        "rate_per_carpet_sqft": (total / carpet) if carpet else None,
        "line_count": len(lines),
        "unit_types": len({l.unit_type_id for l in lines}),
    }


def usage(model: ProjectModel, params: ParameterSet, kind: str,
          subject: str) -> dict[str, Any]:
    """Where a value is used -- the question provenance cannot answer.

    "If I change this, what moves?" The workbook has no way to ask it: 10.764
    is typed into hundreds of formulas with nothing linking them, which is why
    nobody dares touch one.
    """
    from qs_engine.rules import usage as U

    finder = {"parameter": U.parameter_usage, "rate": U.rate_usage,
              "room": U.room_usage}.get(kind)
    if finder is None:
        raise KeyError(f"cannot look up usage of a {kind!r}")

    found = finder(subject, model, params)
    return {
        "subject": found.subject, "kind": found.kind,
        "description": found.description, "note": found.note,
        "total_amount": found.total_amount, "total_lines": found.total_lines,
        "uses": [{"where": u.where, "detail": u.detail, "quantity": u.quantity,
                  "unit": u.unit, "amount": u.amount} for u in found.uses],
    }


# --------------------------------------------------------------------------
# Cost lines and the project roll-up
# --------------------------------------------------------------------------

def _priced_line_json(p) -> dict[str, Any]:
    return {
        "id": p.line.id, "section_id": p.line.section_id,
        "description": p.description, "unit": p.unit, "qty": p.qty,
        "rate": p.rate, "amount": p.amount, "status": p.status,
        "depth": p.depth, "is_heading": p.is_heading, "message": p.message,
        "source_ref": p.line.source_ref, "qty_carried": p.line.qty_carried,
        "exclusion_reason": p.line.exclusion_reason,
        "rate_item_id": p.line.rate_item_id, "manual_rate": p.line.manual_rate,
        "qty_derivation": _derivation(p.qty_derivation) if p.qty_derivation else None,
        "rate_derivation": _derivation(p.rate_derivation) if p.rate_derivation else None,
    }


def cost_lines(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """Every cost line, priced, grouped by the section that owns it."""
    from qs_engine.rules.cost_lines import (compute_cost_lines, excluded,
                                            section_total, total_cost, unpriced)

    lines = compute_cost_lines(model, params)
    return {
        "sections": [{
            "id": s.id, "code": s.code, "name": s.name, "seq": s.seq,
            "excel_ref": s.excel_ref,
            "amount": section_total(lines, s.id),
            "lines": [_priced_line_json(l) for l in lines
                      if l.line.section_id == s.id],
        } for s in sorted(model.cost_sections, key=lambda x: (x.seq, x.code))],
        "total": total_cost(lines),
        "excluded": [_priced_line_json(l) for l in excluded(lines)],
        "unpriced": [_priced_line_json(l) for l in unpriced(lines)],
    }


def project_summary(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """The number at the bottom, and everything that makes it."""
    from qs_engine.rules.summary import project_summary as compute

    # The workbook divides by construction area (Construction Area!S45), not
    # carpet area -- they differ by about 2.5x, so the wrong one here would
    # quietly report a rate that looks plausible and is not.
    s = compute(model, params, params["construction_area_sqft"])
    return {
        "sections": [{"id": x.id, "code": x.code, "name": x.name,
                      "amount": x.amount, "lines": x.lines,
                      "carried": x.carried, "is_carried": x.is_carried,
                      "excel_ref": x.excel_ref,
                      "derivation": _derivation(x.derivation)
                      if x.derivation else None} for x in s.sections],
        "subtotal": s.subtotal,
        "uplifts": [{"code": u.code, "label": u.label, "rate": u.rate,
                     "amount": u.amount, "basis": u.basis} for u in s.uplifts],
        "before_tax": s.before_tax, "tax": s.tax, "total": s.total,
        "construction_area_sqft": s.construction_area_sqft,
        "rate_per_sqft": s.rate_per_sqft,
        "derivation": _derivation(s.derivation) if s.derivation else None,
    }


# --------------------------------------------------------------------------
# The project dashboard, and the copy that leaves the building
# --------------------------------------------------------------------------

def project_card(model: ProjectModel, params: ParameterSet) -> dict[str, Any]:
    """Enough of a project's shape to choose between them on a dashboard."""
    from qs_engine.rules.cost_lines import compute_cost_lines, total_cost

    report = validate(model, params)
    return {
        "floors": len(model.floors),
        "unit_types": len(model.unit_types),
        "rooms": len(model.unit_type_rooms),
        "units": sum(model.unit_count(u.id) for u in model.unit_types
                     if not u.is_common_area),
        "cost_lines": len(model.cost_lines),
        "cost_total": total_cost(compute_cost_lines(model, params)),
        "health": round(report.health_score()),
        "blocking": len(report.blocking),
        "can_issue": report.can_issue,
    }


#: How a revision suffix is written: "R" and a number, at the end of a name.
_REVISION = re.compile(r"^(?P<base>.*?)\s*R(?P<number>\d+)\s*$", re.IGNORECASE)


def revision_base(name: str) -> str:
    """The name with any revision suffix taken off.

    ``AVS Rudraksh R1`` and ``AVS Rudraksh`` are revisions of one estimate, so
    copying either offers the next number in the same series rather than
    starting a second one.
    """
    match = _REVISION.match(name.strip())
    return (match.group("base") if match else name.strip()) or name.strip()


def next_revision_name(name: str, taken: Iterable[str]) -> str:
    """The next unused revision of ``name``.

    Two copies of the same project both offered "R1" and both took it, which is
    how an installation ends up with two projects of one name and no way to
    tell them apart.  The next number is read off the names already in use, so
    the second copy is R2 because R1 exists -- not because anything counted the
    clicks.
    """
    base = revision_base(name)
    folded = {n.strip().casefold() for n in taken}
    used = set()
    for existing in folded:
        match = _REVISION.match(existing)
        if match and match.group("base").strip() == base.casefold():
            used.add(int(match.group("number")))
    # Revisions go forward.  If R7 is the highest, the next is R8 even when R1
    # through R6 were never made -- filling a gap would give the new copy a
    # number that reads as older than the thing it was copied from.
    number = max(used) + 1 if used else 1
    while f"{base} r{number}".casefold() in folded:
        number += 1
    return f"{base} R{number}"


def new_project(name: str, city: str = "", client: str = "") -> ProjectModel:
    """An empty estimate.

    Everything a project holds is a list, and they all start empty: a new
    estimate is not a copy of a template with someone else's rooms in it.
    """
    import re as _re
    import uuid

    slug = _re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or "project"
    return ProjectModel(project=Project(
        id=f"{slug}-{uuid.uuid4().hex[:6]}", code=name[:24].upper(),
        name=name, city=city, client=client))


def duplicate(model: ProjectModel, name: str) -> ProjectModel:
    """Copy a project so the two can never share a row.

    Every id is rewritten, and every reference is repointed at the new one, so
    editing the copy cannot reach back into the original. A shallow copy would
    look right on screen and quietly write through.
    """
    import copy
    import dataclasses
    import re
    import uuid

    from qs_engine import model as M

    fresh = copy.deepcopy(model)
    suffix = uuid.uuid4().hex[:6]
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or "project"
    new_project_id = f"{slug}-{suffix}"

    # Old id -> new id, for every record in the model.
    mapping: dict[str, str] = {model.project.id: new_project_id}
    collections = [f.name for f in dataclasses.fields(fresh)
                   if f.name != "project"]
    for attr in collections:
        for item in getattr(fresh, attr):
            if getattr(item, "id", None):
                mapping[item.id] = f"{item.id}-{suffix}"

    def repoint(item) -> None:
        for field in dataclasses.fields(item):
            value = getattr(item, field.name)
            if isinstance(value, str) and value in mapping:
                setattr(item, field.name, mapping[value])

    fresh.project = M.Project(id=new_project_id, code=name[:24], name=name,
                              city=model.project.city, client=model.project.client)
    for attr in collections:
        for item in getattr(fresh, attr):
            repoint(item)
    return fresh


#: The sheets the export carries, in the order a reader opens them.
EXPORT_SHEETS = ("Summary", "Cost Lines", "Take-off", "Rate Library",
                 "Rooms", "Openings", "Parameters", "Reconciliation")


def export_workbook(model: ProjectModel, params: ParameterSet) -> bytes:
    """The estimate as a workbook.

    Written with formulas rather than values wherever a figure is derived, so a
    reader can see how it was reached and check it in Excel. That is the
    opposite of what this platform replaced -- but an exported file is read by
    people who do not have the platform, and a number they cannot check is a
    number they cannot trust.
    """
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from qs_engine.rules.cost_lines import compute_cost_lines
    from qs_engine.rules.summary import project_summary
    from qs_engine.rules.takeoff import by_finish, compute_takeoff

    head = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")
    money = '#,##0.00'
    title = Font(bold=True, size=12)

    wb = Workbook()
    wb.remove(wb.active)

    def sheet(name: str, columns: list[tuple[str, int]]):
        ws = wb.create_sheet(name[:31])
        for i, (label, width) in enumerate(columns, start=1):
            cell = ws.cell(1, i, label)
            cell.font, cell.fill = head, head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        return ws

    # -- Summary -----------------------------------------------------------
    summary = project_summary(model, params, params["construction_area_sqft"])
    ws = sheet("Summary", [("Section", 30), ("Lines", 10), ("Quantities", 16),
                           ("Amount", 20)])
    ws.cell(1, 1).value = "Section"
    row = 2
    first = row
    for section in summary.sections:
        ws.cell(row, 1, section.name)
        ws.cell(row, 2, section.lines)
        ws.cell(row, 3, "carried" if section.is_carried else "derived here")
        ws.cell(row, 4, section.amount).number_format = money
        row += 1
    last = row - 1

    ws.cell(row, 1, "Subtotal").font = title
    ws.cell(row, 4, f"=SUM(D{first}:D{last})").number_format = money
    subtotal_row = row
    row += 1
    for uplift in summary.uplifts:
        ws.cell(row, 1, f"{uplift.label} @ {uplift.rate:.0%}")
        ws.cell(row, 4, f"=D{subtotal_row}*{uplift.rate}").number_format = money
        row += 1
    ws.cell(row, 1, "Before tax").font = title
    ws.cell(row, 4, f"=SUM(D{subtotal_row}:D{row - 1})").number_format = money
    before_tax = row
    row += 1
    gst = params["gst_pct"]
    ws.cell(row, 1, f"GST @ {gst:.0%}")
    ws.cell(row, 4, f"=D{before_tax}*{gst}").number_format = money
    row += 1
    ws.cell(row, 1, "Project total").font = title
    ws.cell(row, 4, f"=D{before_tax}+D{row - 1}").number_format = money
    ws.cell(row, 4).font = title
    row += 2
    ws.cell(row, 1, "Exported from the DBOT QS Platform. Section totals are "
                    "filters over the lines that belong to them, so a band "
                    "cannot outgrow the range that sums it.")

    # -- Cost lines --------------------------------------------------------
    ws = sheet("Cost Lines", [("Section", 24), ("Description", 44), ("Unit", 8),
                              ("Quantity", 14), ("Rate", 14), ("Amount", 18),
                              ("From", 26)])
    sections = {s.id: s.name for s in model.cost_sections}
    row = 2
    for line in compute_cost_lines(model, params):
        if line.is_heading:
            ws.cell(row, 2, line.description).font = Font(bold=True)
            row += 1
            continue
        ws.cell(row, 1, sections.get(line.line.section_id, ""))
        ws.cell(row, 2, line.description)
        ws.cell(row, 3, line.unit)
        ws.cell(row, 4, line.qty)
        ws.cell(row, 5, line.rate)
        ws.cell(row, 6, f"=D{row}*E{row}").number_format = money
        ws.cell(row, 7, line.line.source_ref)
        row += 1

    # -- Take-off ----------------------------------------------------------
    lines = compute_takeoff(model, params)
    ws = sheet("Take-off", [("Unit type", 18), ("Room", 24), ("Finish", 24),
                            ("Rule", 16), ("Gross", 12), ("Deduction", 12),
                            ("Net", 12), ("Unit", 8), ("Units", 8),
                            ("Total qty", 14), ("Rate", 12), ("Amount", 16)])
    row = 2
    for line in lines:
        ws.cell(row, 1, line.unit_type_code)
        ws.cell(row, 2, line.room_label)
        ws.cell(row, 3, line.finish_name)
        ws.cell(row, 4, line.qty_rule)
        ws.cell(row, 5, line.gross)
        ws.cell(row, 6, line.deduction)
        ws.cell(row, 7, f"=E{row}-F{row}")
        ws.cell(row, 8, line.unit)
        ws.cell(row, 9, line.unit_count)
        ws.cell(row, 10, f"=G{row}*I{row}")
        ws.cell(row, 11, line.rate)
        ws.cell(row, 12, f"=J{row}*K{row}").number_format = money
        row += 1

    ws = sheet("Take-off by finish", [("Finish", 28), ("Quantity", 14),
                                      ("Unit", 8), ("Sq ft", 14),
                                      ("Amount", 18)])
    row = 2
    for group in by_finish(lines, params):
        ws.cell(row, 1, group.label)
        ws.cell(row, 2, group.quantity)
        ws.cell(row, 3, group.unit)
        ws.cell(row, 4, group.quantity_sqft)
        ws.cell(row, 5, group.amount).number_format = money
        row += 1

    # -- Rate library ------------------------------------------------------
    ws = sheet("Rate Library", [("Code", 16), ("Description", 34),
                                ("Specification", 30), ("Unit", 8),
                                ("Method", 20), ("Basic", 12), ("Laying", 12),
                                ("Wastage", 10), ("Overall rate", 14)])
    row = 2
    for item in model.rate_items:
        revision = model.current_revision(item.id)
        ws.cell(row, 1, item.code)
        ws.cell(row, 2, item.description)
        ws.cell(row, 3, item.specification)
        ws.cell(row, 4, item.unit)
        ws.cell(row, 5, revision.method.value if revision else "")
        ws.cell(row, 6, revision.basic_rate if revision else None)
        ws.cell(row, 7, revision.laying_rate if revision else None)
        ws.cell(row, 8, revision.wastage_pct if revision else None)
        try:
            ws.cell(row, 9, effective_rate(item, model, params).value)
        except RateBuildupError:
            ws.cell(row, 9, None)
        row += 1

    # -- Rooms -------------------------------------------------------------
    ws = sheet("Rooms", [("Unit type", 18), ("Room", 26), ("Room type", 24),
                         ("Nos", 8), ("Area sq.m", 12), ("Perimeter m", 12),
                         ("Area sq.ft", 14)])
    factor = params["factor_sqm_to_sqft"]
    row = 2
    for unit in sorted(model.unit_types, key=lambda u: u.seq):
        for room in model.rooms_of(unit.id):
            ws.cell(row, 1, unit.code)
            ws.cell(row, 2, room.label)
            ws.cell(row, 3, model.room_type(room.room_type_id).name)
            ws.cell(row, 4, room.count_per_unit)
            ws.cell(row, 5, room.carpet_area_sqm)
            ws.cell(row, 6, room.perimeter_m)
            ws.cell(row, 7, f"=E{row}*{factor}")
            row += 1

    # -- Openings ----------------------------------------------------------
    from qs_engine.rules.schedule import priced_opening_schedule
    ws = sheet("Openings", [("Code", 14), ("Kind", 14), ("Width m", 10),
                            ("Height m", 10), ("Count", 10), ("Quantity", 14),
                            ("Unit", 8), ("Rate", 14), ("Amount", 18)])
    row = 2
    for line in priced_opening_schedule(model, params):
        ws.cell(row, 1, line.code)
        ws.cell(row, 2, line.kind.value)
        ws.cell(row, 3, line.width_m)
        ws.cell(row, 4, line.height_m)
        ws.cell(row, 5, line.count)
        ws.cell(row, 6, line.quantity)
        ws.cell(row, 7, line.unit)
        ws.cell(row, 8, line.rate)
        ws.cell(row, 9, line.amount).number_format = money
        row += 1

    # -- Parameters --------------------------------------------------------
    ws = sheet("Parameters", [("Key", 26), ("Value", 14), ("Unit", 12),
                              ("What it is", 60), ("From", 40)])
    row = 2
    for parameter in params:
        ws.cell(row, 1, parameter.key)
        ws.cell(row, 2, parameter.value)
        ws.cell(row, 3, parameter.unit)
        ws.cell(row, 4, parameter.description or "— not yet described —")
        ws.cell(row, 5, parameter.source)
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
