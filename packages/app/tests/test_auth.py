"""Accounts, roles, the dashboard and the export.

The audit log recorded every write since the store was written, and every row
said "local". A change log that cannot name a person is a list of events, not
an account of what happened.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient  # noqa: E402

from qs_app import auth  # noqa: E402


@pytest.fixture
def app_client(tmp_path, avs):
    """A fresh installation per test -- accounts change how the app behaves."""
    from qs_app import server
    from qs_app.store import Store

    db = tmp_path / "auth.db"
    store = Store(db)
    store.save(avs.model, avs.params)

    server.state.store = store
    server.state.open(avs.model.project.id)
    with TestClient(server.app) as client:
        yield client


# -- passwords -------------------------------------------------------------

def test_a_password_is_hashed_not_stored():
    stored = auth.hash_password("a-good-password")
    assert "a-good-password" not in stored
    assert stored.startswith("scrypt$")
    assert auth.verify_password("a-good-password", stored)
    assert not auth.verify_password("a-good-Password", stored)


def test_a_short_password_is_refused():
    with pytest.raises(ValueError):
        auth.hash_password("short")


def test_an_unreadable_hash_is_a_refusal_not_a_crash():
    assert not auth.verify_password("anything", "not-a-hash")
    assert not auth.verify_password("anything", "")


def test_two_identical_passwords_hash_differently():
    """Salted, so a stolen table cannot be matched against itself."""
    assert auth.hash_password("same-password") != auth.hash_password("same-password")


# -- the first account -----------------------------------------------------

def test_the_platform_is_open_until_somebody_creates_an_account(app_client):
    """A fresh clone runs with `make run` and no ceremony."""
    me = app_client.get("/api/me").json()
    assert me["open_access"] and not me["signed_in"]

    rooms = app_client.get("/api/unit-types").json()
    unit = next(u for u in rooms if u["rooms"])
    detail = app_client.get(f"/api/unit-types/{unit['id']}/rooms").json()
    room = detail["rooms"][0]
    assert app_client.patch(f"/api/collections/rooms/{room['id']}",
                            json={"label": "Renamed while open"}).status_code == 200


@pytest.fixture
def accounts_required():
    """Turn the sign-in gate on for one test.

    It is off by default (``server.ACCOUNTS_REQUIRED``), at the user's request.
    The accounts underneath it are kept, so what guards them is kept too --
    switched on here rather than deleted, which is what tells us the flag is
    the only thing standing between the platform and a working sign-in.
    """
    from qs_app import server
    server.ACCOUNTS_REQUIRED = True
    try:
        yield
    finally:
        server.ACCOUNTS_REQUIRED = False


def test_the_gate_is_off_so_nobody_is_asked_to_sign_in(app_client):
    """What "remove the sign in feature" means, asserted.

    Not that the accounts are gone -- that a write goes through with nobody
    signed in and nothing asked for.
    """
    me = app_client.get("/api/me").json()
    assert me["accounts_required"] is False
    assert me["open_access"] is True

    app_client.post("/api/users", json={"name": "Aditya", "email": "a@b.com",
                                        "password": "a-good-password"})
    unit = next(u for u in app_client.get("/api/unit-types").json() if u["rooms"])
    room = app_client.get(f"/api/unit-types/{unit['id']}/rooms").json()["rooms"][0]
    written = app_client.patch(f"/api/collections/rooms/{room['id']}",
                               json={"label": "Went straight through"})
    assert written.status_code == 200, "an account exists and the write still goes"


def test_the_first_account_closes_it_and_is_an_admin(app_client, accounts_required):
    created = app_client.post("/api/users", json={
        "name": "Aditya", "email": "aditya@dbotrealty.com",
        "password": "a-good-password"}).json()
    assert created["first_user"]
    assert created["user"]["role"] == "admin", "somebody has to be able to administer"

    me = app_client.get("/api/me").json()
    assert not me["open_access"]

    # Writing now needs a sign-in.
    rooms = app_client.get("/api/unit-types").json()
    unit = next(u for u in rooms if u["rooms"])
    room = app_client.get(f"/api/unit-types/{unit['id']}/rooms").json()["rooms"][0]
    assert app_client.patch(f"/api/collections/rooms/{room['id']}",
                            json={"label": "Nope"}).status_code == 401


def test_a_second_account_needs_an_admin(app_client):
    app_client.post("/api/users", json={"name": "Aditya", "email": "a@b.com",
                                        "password": "a-good-password"})
    assert app_client.post("/api/users", json={
        "name": "Someone", "email": "c@d.com", "password": "another-password",
        "role": "admin"}).status_code == 403


# -- signing in ------------------------------------------------------------

def _sign_up_and_in(client, role="admin", email="a@b.com"):
    client.post("/api/users", json={"name": "Aditya", "email": "a@b.com",
                                    "password": "a-good-password"})
    client.post("/api/login", json={"email": "a@b.com",
                                    "password": "a-good-password"})
    if role != "admin":
        client.post("/api/users", json={"name": "Ravi", "email": email,
                                        "password": "another-password",
                                        "role": role})
        client.post("/api/logout")
        client.post("/api/login", json={"email": email,
                                        "password": "another-password"})


def test_a_wrong_password_says_the_same_thing_as_a_wrong_email(app_client):
    """So the message cannot be used to discover which addresses have accounts."""
    app_client.post("/api/users", json={"name": "Aditya", "email": "a@b.com",
                                        "password": "a-good-password"})
    wrong_password = app_client.post("/api/login", json={
        "email": "a@b.com", "password": "guessing"})
    no_such_user = app_client.post("/api/login", json={
        "email": "nobody@b.com", "password": "guessing"})

    assert wrong_password.status_code == no_such_user.status_code == 401
    assert wrong_password.json()["detail"] == no_such_user.json()["detail"]


def test_signing_in_lets_the_write_through_and_names_the_person(app_client):
    _sign_up_and_in(app_client)

    unit = next(u for u in app_client.get("/api/unit-types").json() if u["rooms"])
    room = app_client.get(f"/api/unit-types/{unit['id']}/rooms").json()["rooms"][0]
    assert app_client.patch(f"/api/collections/rooms/{room['id']}",
                            json={"label": "Named change"}).status_code == 200

    entries = app_client.get("/api/audit").json()
    assert entries, "the change was logged"
    assert "Aditya" in entries[0]["actor"], \
        f"the log must name the person, not {entries[0]['actor']!r}"
    assert entries[0]["actor"] != "local"


def test_a_reviewer_may_read_but_not_write(app_client, accounts_required):
    """A reviewer's job is to disagree with a number, not to replace it."""
    _sign_up_and_in(app_client, role="reviewer", email="ravi@b.com")

    assert app_client.get("/api/takeoff").status_code == 200

    unit = next(u for u in app_client.get("/api/unit-types").json() if u["rooms"])
    room = app_client.get(f"/api/unit-types/{unit['id']}/rooms").json()["rooms"][0]
    refused = app_client.patch(f"/api/collections/rooms/{room['id']}",
                               json={"label": "Nope"})
    assert refused.status_code == 403
    assert "reviewer" in refused.json()["detail"]


def test_signing_out_ends_the_session(app_client):
    _sign_up_and_in(app_client)
    assert app_client.get("/api/me").json()["signed_in"]
    app_client.post("/api/logout")
    assert not app_client.get("/api/me").json()["signed_in"]


# -- the dashboard ---------------------------------------------------------

def test_the_dashboard_describes_each_project(app_client):
    projects = app_client.get("/api/dashboard").json()["projects"]
    assert projects
    card = projects[0]
    for key in ("units", "rooms", "cost_total", "health", "blocking", "open"):
        assert key in card
    assert card["open"], "the seeded project is the open one"


def test_a_copy_shares_no_rows_with_its_original(app_client):
    """Editing a copy must not reach back into what it came from."""
    from qs_app import server

    original = server.state.store.load("avs")
    made = app_client.post("/api/projects/duplicate",
                           json={"project_id": "avs", "name": "AVS R1"}).json()
    assert made["ok"]

    copy = server.state.store.load(made["project_id"])
    assert copy.project.id != original.project.id
    assert len(copy.unit_type_rooms) == len(original.unit_type_rooms)
    assert not ({r.id for r in copy.unit_type_rooms}
                & {r.id for r in original.unit_type_rooms})
    assert not ({r.id for r in copy.rate_items}
                & {r.id for r in original.rate_items})


def test_editing_a_copy_leaves_the_original_exactly_as_it_was(app_client):
    """R1 is where the work happens; R0 must not move a figure.

    Id-disjointness is necessary and not sufficient -- a copy could share a
    parameter set, or write through a table the duplicate forgot. This reads
    every row of the original back after editing the copy and compares it
    field by field.
    """
    import dataclasses

    from qs_app import server

    def snapshot(project_id):
        model = server.state.store.load(project_id)
        params = server.state.store.load_params(project_id)
        return ({attr: [dataclasses.asdict(i) for i in getattr(model, attr)]
                 for attr in (f.name for f in dataclasses.fields(model))
                 if isinstance(getattr(model, attr), list)},
                params.as_dict())

    before = snapshot("avs")
    made = app_client.post("/api/projects/duplicate",
                           json={"project_id": "avs"}).json()
    copy_id = made["project_id"]

    copy = server.state.store.load(copy_id)
    room = copy.unit_type_rooms[0]
    room.carpet_area_sqm = room.carpet_area_sqm + 99.0
    room.label = "edited in the copy"
    copy.unit_type_rooms.pop()
    copy.project.city = "Somewhere else"
    server.state.store.save(copy, server.state.store.load_params(copy_id))

    assert snapshot("avs") == before, "editing the copy changed the original"
    assert server.state.store.load("avs").project.city != "Somewhere else"


def test_a_second_copy_is_r2_not_a_second_r1(app_client):
    """Two projects of one name cannot be told apart on any screen."""
    first = app_client.post("/api/projects/duplicate",
                            json={"project_id": "avs"}).json()
    second = app_client.post("/api/projects/duplicate",
                             json={"project_id": "avs"}).json()
    names = {p["id"]: p["name"]
             for p in app_client.get("/api/dashboard").json()["projects"]}
    assert names[first["project_id"]] != names[second["project_id"]]
    assert names[second["project_id"]].endswith("R2")

    # And the screen is told the name rather than guessing it.
    cards = {p["id"]: p for p in app_client.get("/api/dashboard").json()["projects"]}
    assert cards["avs"]["next_revision"].endswith("R3")


def test_a_name_already_in_use_is_refused_rather_than_duplicated(app_client):
    made = app_client.post("/api/projects/duplicate",
                           json={"project_id": "avs", "name": "Taken"})
    assert made.status_code == 200
    again = app_client.post("/api/projects/duplicate",
                            json={"project_id": "avs", "name": "taken"})
    assert again.status_code == 409
    assert "already exists" in again.json()["detail"]


def test_a_new_project_starts_empty(app_client):
    from qs_app import server

    made = app_client.post("/api/projects/new",
                           json={"name": "Palm Grove", "city": "Thane"})
    assert made.status_code == 200, made.text
    model = server.state.store.load(made.json()["project_id"])
    assert model.project.name == "Palm Grove" and model.project.city == "Thane"
    assert not model.unit_types and not model.floors and not model.rate_items
    assert not model.unit_type_rooms, "a new estimate is nobody else's rooms"

    cards = {p["id"]: p for p in app_client.get("/api/dashboard").json()["projects"]}
    assert cards[made.json()["project_id"]]["rooms"] == 0

    app_client.post("/api/projects/open", json={"project_id": "avs"})


def test_a_new_project_will_not_take_a_name_in_use(app_client):
    app_client.post("/api/projects/new", json={"name": "Twice Over"})
    again = app_client.post("/api/projects/new", json={"name": "twice over"})
    assert again.status_code == 409
    app_client.post("/api/projects/open", json={"project_id": "avs"})


def test_archiving_keeps_the_project(app_client):
    app_client.post("/api/projects/archive",
                    json={"project_id": "avs", "archived": True})
    projects = {p["id"]: p for p in app_client.get("/api/dashboard").json()["projects"]}
    assert projects["avs"]["archived"]
    assert projects["avs"]["rooms"] > 0, "archived, not deleted"


def test_an_archived_project_comes_back_whole(app_client):
    """Archiving is reversible or it is deletion with a nicer word."""
    rooms = {p["id"]: p["rooms"]
             for p in app_client.get("/api/dashboard").json()["projects"]}["avs"]

    app_client.post("/api/projects/archive",
                    json={"project_id": "avs", "archived": True})
    restored = app_client.post("/api/projects/restore", json={"project_id": "avs"})
    assert restored.status_code == 200
    assert restored.json()["archived"] is False

    card = {p["id"]: p
            for p in app_client.get("/api/dashboard").json()["projects"]}["avs"]
    assert not card["archived"]
    assert card["rooms"] == rooms, "came back with every row it went in with"


def test_restoring_a_project_that_is_not_there_says_so(app_client):
    missing = app_client.post("/api/projects/restore", json={"project_id": "nope"})
    assert missing.status_code == 404


# -- the export ------------------------------------------------------------

def test_the_export_is_a_readable_workbook(app_client):
    import io

    import openpyxl

    response = app_client.get("/api/export.xlsx")
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert {"Summary", "Cost Lines", "Take-off", "Rate Library",
            "Rooms", "Openings", "Parameters"} <= set(wb.sheetnames)
    assert wb["Take-off"].max_row > 1000
    assert wb["Cost Lines"].max_row > 100


def test_the_export_carries_formulas_so_a_reader_can_check_it(app_client):
    """The file leaves the building; a number nobody can check is not trusted."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(app_client.get("/api/export.xlsx").content))
    take_off = wb["Take-off"]
    assert str(take_off.cell(2, 12).value).startswith("="), \
        "the amount column should be a formula, not a pasted number"
    assert str(wb["Summary"].cell(2 + 9, 4).value).startswith("=")
